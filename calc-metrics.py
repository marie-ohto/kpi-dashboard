#!/usr/bin/env python3
"""
KPIダッシュボード用メトリクス計算スクリプト

Usage:
    python calc-metrics.py <excel_path> [--output metrics.json]
                           [--sheet-work 作業時間と処理数]
                           [--sheet-volume 投入量]
                           [--sheet-staff 人員表]
                           [--sheet-baseline 基準値]
                           [--work-hours 7]

Output:
    メトリクスJSON — スキーマ仕様は references/views-spec.md（メトリクスJSON仕様）を参照。
"""

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl で導入してください。")
    sys.exit(1)


def parse_args():
    p = argparse.ArgumentParser(description="KPIダッシュボード用メトリクス計算")
    p.add_argument("excel_path", help="Excelファイルパス")
    p.add_argument("--output", default="/tmp/dashboard-metrics.json", help="出力JSONパス")
    p.add_argument("--sheet-work", default="作業時間と処理数")
    p.add_argument("--sheet-volume", default="投入量")
    p.add_argument("--sheet-staff", default="人員表")
    p.add_argument("--sheet-baseline", default="基準値")
    p.add_argument("--work-hours", type=float, default=7.0,
                   help="1日の平均稼働時間（人員過不足指数の適正ライン計算に使用）")
    return p.parse_args()


def load_workbook(path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: ファイルが見つかりません: {path}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Excelファイルの読み込みに失敗しました: {e}")
        sys.exit(1)


def get_sheet(wb, name):
    if name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        print(f"ERROR: シート '{name}' が見つかりません。")
        print(f"  利用可能なシート: {available}")
        print(f"  --sheet-* オプションでシート名を指定してください。")
        sys.exit(1)
    return wb[name]


def sheet_to_rows(sheet):
    """シートを [{列名: 値, ...}] のリストに変換する"""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def find_col(row, candidates):
    """複数の候補列名からヒットした最初のキーを返す"""
    for c in candidates:
        if c in row:
            return c
    return None


def to_date_str(val):
    """日付値を YYYY-MM-DD 文字列に変換する"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return str(val)


def safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


def calculate(args):
    wb = load_workbook(args.excel_path)

    # シートを読み込む
    work_rows = sheet_to_rows(get_sheet(wb, args.sheet_work))
    volume_rows = sheet_to_rows(get_sheet(wb, args.sheet_volume))
    staff_rows = sheet_to_rows(get_sheet(wb, args.sheet_staff))
    baseline_rows = sheet_to_rows(get_sheet(wb, args.sheet_baseline))

    # 基準値を取得
    baseline_col = find_col(baseline_rows[0] if baseline_rows else {}, ["基準値", "基準値(件/h)", "基準"])
    baseline_productivity = safe_float(
        baseline_rows[0].get(baseline_col) if baseline_rows and baseline_col else None,
        default=34.0
    )

    # 列名を動的に解決する
    sample_work = work_rows[0] if work_rows else {}
    col_bureau = find_col(sample_work, ["局名", "支社名", "拠点名", "部署名"])
    col_name = find_col(sample_work, ["氏名", "名前", "スタッフ名"])
    col_date = find_col(sample_work, ["日付", "作業日", "日時"])
    col_hours = find_col(sample_work, ["作業時間", "作業時間(h)", "稼働時間", "時間"])
    col_count = find_col(sample_work, ["処理件数", "件数", "処理数"])

    missing = [name for name, col in [
        ("局名/支社名", col_bureau), ("氏名", col_name),
        ("日付", col_date), ("作業時間", col_hours), ("処理件数", col_count)
    ] if col is None]
    if missing:
        print(f"ERROR: 作業時間シートに必要な列が見つかりません: {', '.join(missing)}")
        print(f"  利用可能な列: {', '.join(sample_work.keys())}")
        sys.exit(1)

    sample_vol = volume_rows[0] if volume_rows else {}
    col_vol_bureau = find_col(sample_vol, ["局名", "支社名", "拠点名"])
    col_vol_date = find_col(sample_vol, ["日付", "作業日"])
    col_vol_in = find_col(sample_vol, ["投入量", "投入件数", "受付件数"])
    col_vol_done = find_col(sample_vol, ["消化量", "消化件数", "完了件数", "処理件数"])

    # 人員表から局別スタッフ一覧を構築
    sample_staff = staff_rows[0] if staff_rows else {}
    col_staff_bureau = find_col(sample_staff, ["局名", "支社名", "拠点名"])
    col_staff_name = find_col(sample_staff, ["氏名", "名前", "スタッフ名"])
    bureau_staff = defaultdict(list)
    for r in staff_rows:
        b = str(r.get(col_staff_bureau, "")).strip()
        n = str(r.get(col_staff_name, "")).strip()
        if b and n:
            bureau_staff[b].append(n)

    # 作業時間シートから個人別・日別の生産性を計算
    # {bureau: {name: {date: productivity}}}
    individual_daily = defaultdict(lambda: defaultdict(dict))
    for r in work_rows:
        bureau = str(r.get(col_bureau, "")).strip()
        name = str(r.get(col_name, "")).strip()
        date = to_date_str(r.get(col_date))
        hours = safe_float(r.get(col_hours))
        count = safe_float(r.get(col_count))
        if not bureau or not name or not date:
            continue
        prod = round(count / hours, 2) if hours > 0 else 0.0
        individual_daily[bureau][name][date] = prod

    # 投入量シートから局別・日別の指標を計算
    # {bureau: {date: {volume, completion_rate}}}
    bureau_volume = defaultdict(lambda: defaultdict(dict))
    for r in volume_rows:
        bureau = str(r.get(col_vol_bureau, "")).strip()
        date = to_date_str(r.get(col_vol_date))
        vol_in = safe_float(r.get(col_vol_in) if col_vol_in else None)
        vol_done = safe_float(r.get(col_vol_done) if col_vol_done else None)
        if not bureau or not date:
            continue
        bureau_volume[bureau][date] = {
            "volume": vol_in,
            "completion_rate": round(vol_done / vol_in, 4) if vol_in > 0 else 0.0
        }

    # 日付範囲を取得
    all_dates = sorted({
        d for bd in bureau_volume.values() for d in bd.keys()
    } | {
        d for bid in individual_daily.values() for nd in bid.values() for d in nd.keys()
    })
    period_start = all_dates[0] if all_dates else ""
    period_end = all_dates[-1] if all_dates else ""

    # 局一覧（投入量 or 作業時間シートに登場した局）
    all_bureaus = sorted(set(bureau_volume.keys()) | set(individual_daily.keys()))

    bureaus_out = []
    for bureau in all_bureaus:
        staff_list = bureau_staff.get(bureau, [])
        staff_count = len(staff_list) if staff_list else max(
            1, len(individual_daily.get(bureau, {}))
        )

        # 局別日別生産性（全スタッフ平均）
        daily_prod = {}
        for date in all_dates:
            day_prods = [
                individual_daily[bureau][n][date]
                for n in individual_daily.get(bureau, {})
                if date in individual_daily[bureau].get(n, {})
            ]
            if day_prods:
                daily_prod[date] = round(sum(day_prods) / len(day_prods), 2)

        avg_prod = round(sum(daily_prod.values()) / len(daily_prod), 2) if daily_prod else 0.0
        standard_ratio = round(avg_prod / baseline_productivity, 4) if baseline_productivity else 0.0

        # 投入量・消化率
        daily_volume = {d: v["volume"] for d, v in bureau_volume.get(bureau, {}).items()}
        daily_completion = {d: v["completion_rate"] for d, v in bureau_volume.get(bureau, {}).items()}

        # 人員過不足指数（latest: 最終日、avg: 全期間平均）
        manpower_index_latest = 0.0
        if period_end and period_end in daily_volume and staff_count > 0:
            manpower_index_latest = round(daily_volume[period_end] / staff_count, 1)

        manpower_index_avg = 0.0
        if daily_volume and staff_count > 0:
            manpower_index_avg = round(sum(daily_volume.values()) / len(daily_volume) / staff_count, 1)

        # 個人別メトリクス
        individuals = []
        for name in sorted(individual_daily.get(bureau, {}).keys()):
            ind_daily = individual_daily[bureau][name]
            ind_avg = round(sum(ind_daily.values()) / len(ind_daily), 2) if ind_daily else 0.0
            ind_ratio = round(ind_avg / baseline_productivity, 4) if baseline_productivity else 0.0
            individuals.append({
                "name": name,
                "avg_productivity": ind_avg,
                "standard_ratio": ind_ratio,
                "daily_productivity": dict(sorted(ind_daily.items()))
            })

        # 生産性の高い順にソート
        individuals.sort(key=lambda x: x["avg_productivity"], reverse=True)

        bureaus_out.append({
            "name": bureau,
            "staff_count": staff_count,
            "avg_productivity": avg_prod,
            "standard_ratio": standard_ratio,
            "daily_volume": dict(sorted(daily_volume.items())),
            "daily_completion_rate": dict(sorted(daily_completion.items())),
            "daily_productivity": dict(sorted(daily_prod.items())),
            "manpower_index_latest": manpower_index_latest,
            "manpower_index_avg": manpower_index_avg,
            "individuals": individuals
        })

    # 生産性の高い順にソート
    bureaus_out.sort(key=lambda x: x["avg_productivity"], reverse=True)

    return {
        "baseline_productivity": baseline_productivity,
        "work_hours_per_day": args.work_hours,
        "period": {
            "start": period_start,
            "end": period_end,
            "days": len(all_dates)
        },
        "bureaus": bureaus_out
    }


def main():
    args = parse_args()
    metrics = calculate(args)

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(metrics, f, ensure_ascii=False, indent=2)

    print(f"✓ メトリクス計算完了: {args.output}")
    print(f"  集計期間: {metrics['period']['start']} — {metrics['period']['end']} ({metrics['period']['days']}日)")
    print(f"  基準値: {metrics['baseline_productivity']}件/h")
    print(f"  支社数: {len(metrics['bureaus'])}")
    for b in metrics["bureaus"]:
        ratio_pct = round(b["standard_ratio"] * 100)
        flag = "⚠️ " if ratio_pct < 90 else "  "
        print(f"  {flag}{b['name']}: {b['avg_productivity']}件/h ({ratio_pct}%) スタッフ{b['staff_count']}名")


if __name__ == "__main__":
    main()
